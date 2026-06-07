# =============================================================================
# app.py — 学生管理系统 Flask 主应用
# =============================================================================
# 功能：基于 Flask + PyMySQL + MySQL 的 B/S 架构学生管理系统
# 角色：管理员(admin) / 教师(teacher) / 学生(student)
# 核心功能：
#   - 统一登录认证 (SHA256密码哈希)，session 管理
#   - 管理员：学院/专业/班级/学生/教师/课程/成绩/奖惩/审批/账号 CRUD
#   - 教师：查看授课课程、录入成绩、查看统计
#   - 学生：查看成绩/GPA/排名、选课/退课、转专业/放弃成绩/奖项申请
#   - 4.3 分制 GPA 计算 (MySQL 函数 + 存储过程)
#   - 学期选课时间窗控制 (自由选退 → 仅退课 → 锁定)
#   - 学院内同年级排名
# =============================================================================
# app.py - 学生管理系统

import hashlib, os
from flask import Flask, render_template, request, redirect, url_for, session, flash, send_from_directory, jsonify
import pymysql
from functools import wraps

app = Flask(__name__)
app.secret_key = 'student-management-v6-2026'

DB = {'host':'localhost','port':3306,'user':'root','password':'123456','database':'student_management','charset':'utf8mb4'}

# ============ 选课时间窗配置 ============
from datetime import date
SEMESTER_START = date(2026,2,23)     # 2025-2026春开学
FREE_DROP_END  = date(2026,3,8)      # 第2周结束 (自由退选截止)
DROP_END       = date(2026,5,3)      # 第10周结束 (退课截止)
SEMESTER_END   = date(2026,6,14)     # 学期结束

def week_of_semester(today=None):
    """返回当前是第几教学周(1-16), 以及选课阶段"""
    d = today or date.today()
    if d < SEMESTER_START: return 0, 'before'
    wk = ((d - SEMESTER_START).days // 7) + 1
    if wk > 16: wk = 16
    if d <= FREE_DROP_END: phase = 'free'       # 自由选退
    elif d <= DROP_END:    phase = 'drop_only'  # 只能退课
    else:                  phase = 'locked'     # 锁定
    return wk, phase

def open_enroll_semesters():
    """返回当前可选课的学期列表"""       # 当前学期: 2025-2026春, 已锁定
    return ['2026夏', '2026-2027秋']    # 夏季学期: 2026夏 (正在开放)
                                        # 秋季学期: 2026-2027秋 (正在开放)   
def can_drop_semester(sem):
    """判断某学期的课程是否还能退课"""
    if sem in ['2026夏', '2026-2027秋']:      # 只有未来学期(夏季/秋季)在自由期内可退, 当前学期已锁定
        return True  # 未来学期现在正在开放, 可自由退选
    return False     # 已过去的学期和当前学期都不可退

# ---------- 数据库连接与查询辅助函数 ----------
def db():
    """创建新的数据库连接（每次调用生成独立连接，适合简单操作）"""
    return pymysql.connect(**DB)

def q(sql, args=None, one=False):
    """执行 SELECT 查询并返回结果。
    one=True 返回单行（用于 COUNT/单条查询），否则返回全部行。
    注意：此函数自动 commit，不适用于需要事务回滚的场景。"""
    c=db(); cur=c.cursor(); cur.execute(sql,args)
    r=cur.fetchone() if one else cur.fetchall(); c.commit(); c.close(); return r

def e(sql, args=None):
    """执行 INSERT/UPDATE/DELETE，自动 commit。
    注意：每次调用打开独立连接，不适合多语句事务。"""
    c=db(); cur=c.cursor(); cur.execute(sql,args); c.commit(); c.close()

# ============ 认证与权限装饰器 (AUTH) ============
def login_required(f):
    """装饰器：要求用户已登录，未登录则跳转到登录页"""
    @wraps(f)
    def d(*a,**k):
        if not session.get('logged_in'): return redirect(url_for('login'))
        return f(*a,**k)
    return d

def role_required(*roles):
    """装饰器：限制只有指定角色可访问，如 @role_required('admin','teacher')"""
    def dec(f):
        @wraps(f)
        def d(*a,**k):
            if session.get('role') not in roles:
                flash('无权限访问','danger'); return redirect(url_for('index'))
            return f(*a,**k)
        return d
    return dec

# ---------- 登录 / 登出 / 修改密码 ----------
@app.route('/login', methods=['GET','POST'])
def login():
    if request.method=='POST':
        u=request.form.get('username','').strip()
        pw=hashlib.sha256(request.form.get('password','').encode()).hexdigest()
        r=q("SELECT account_id,role,ref_id FROM account WHERE username=%s AND password_hash=%s",(u,pw),one=True)
        if r:
            session['logged_in']=True; session['username']=u; session['role']=r[1]; session['ref_id']=r[2]
            flash(f'登录成功！角色: {r[1]}','success'); return redirect(url_for('index'))
        flash('用户名或密码错误','danger')
    return render_template('login.html')

@app.route('/logout')
def logout():
    session.clear(); flash('已退出','info'); return redirect(url_for('login'))

@app.route('/change_password',methods=['POST'])
@login_required
def change_password():
    old=request.form.get('old_password','')
    new=request.form.get('new_password','')
    confirm=request.form.get('confirm_password','')
    if not old or not new:
        flash('请填写所有密码字段','danger')
    elif new != confirm:
        flash('两次新密码不一致','danger')
    elif len(new)<4:
        flash('新密码至少4位','danger')
    else:
        old_hash=hashlib.sha256(old.encode()).hexdigest()
        acc=q("SELECT account_id FROM account WHERE username=%s AND password_hash=%s",(session['username'],old_hash),one=True)
        if not acc:
            flash('原密码错误','danger')
        else:
            new_hash=hashlib.sha256(new.encode()).hexdigest()
            e("UPDATE account SET password_hash=%s WHERE account_id=%s",(new_hash,acc[0]))
            flash('密码修改成功','success')
    return redirect(request.referrer or url_for('index'))

# ============ 首页仪表盘（按角色渲染不同视图）============
# 管理员 → admin_dashboard.html  /  教师 → teacher_dashboard.html  /  学生 → student_dashboard.html
@app.route('/')
@login_required
def index():
    role=session['role']
    if role=='admin':
        stats={}
        for t in ['college','major','class','student','teacher','course','sc','account','admin']:
            r=q(f"SELECT COUNT(*) FROM {t}",one=True); stats[t]=r[0] if r else 0
        r=q("SELECT COUNT(*) FROM sc WHERE score IS NOT NULL",one=True); stats['scored']=r[0] if r else 0
        r=q("SELECT COUNT(*) FROM major_transfer WHERE status='pending'",one=True); stats['pending_transfer']=r[0] if r else 0
        r=q("SELECT COUNT(*) FROM score_abandon WHERE status='pending'",one=True); stats['pending_abandon']=r[0] if r else 0
        return render_template('admin_dashboard.html',stats=stats)
    elif role=='teacher':
        tno=session['ref_id']
        courses=q("SELECT c.*, (SELECT COUNT(*) FROM sc WHERE sc.cno=c.cno AND sc.status='normal') as cnt, (SELECT COUNT(*) FROM sc WHERE sc.cno=c.cno AND sc.status='normal' AND sc.score IS NULL) as ungraded FROM course c WHERE c.tno=%s ORDER BY c.semester DESC, c.cno",(tno,))
        return render_template('teacher_dashboard.html',courses=courses,tno=tno)
    else: # student
        sno=session['username']
        stu=q("SELECT s.*,c.name as cn,c.major_id,c.enroll_year FROM student s LEFT JOIN class c ON s.class_id=c.class_id WHERE s.sno=%s",(sno,),one=True)
        # 调用存储过程 sp_student_gpa_rank（GPA + 均分 + 学分 + 学院总人数 + 排名）
        c=db(); cur=c.cursor()
        cur.callproc('sp_student_gpa_rank',(sno,0,0,0,0,0,0))
        cur.execute("SELECT @_sp_student_gpa_rank_1, @_sp_student_gpa_rank_2, @_sp_student_gpa_rank_3, @_sp_student_gpa_rank_4, @_sp_student_gpa_rank_5, @_sp_student_gpa_rank_6")
        gpa_r=cur.fetchone()
        gpa_val=gpa_r[0]; avg_score=gpa_r[1]; avg_score_w=gpa_r[2]; total_credits=gpa_r[3]
        college_stats={'total':gpa_r[4] or 0,'rank':gpa_r[5] or '-'}
        # 专业排名（同专业同年级）
        cur.callproc('sp_student_major_rank',(sno,0,0))
        cur.execute("SELECT @_sp_student_major_rank_1, @_sp_student_major_rank_2")
        mr=cur.fetchone(); c.close()
        major_stats={'total':mr[0] or 0,'rank':mr[1] or '-'}
        scores=q("SELECT sc.cno,c.name,c.credit,sc.score,sc.status,fn_score_to_gp_43(sc.score) as gp FROM sc JOIN course c ON sc.cno=c.cno WHERE sc.sno=%s ORDER BY c.name",(sno,))
        rps=q("SELECT rp.*,ad.name as creator_name FROM reward_punishment rp LEFT JOIN admin ad ON rp.created_by=ad.admin_id WHERE rp.sno=%s ORDER BY rp.rp_date DESC",(sno,))
        transfers=q("SELECT mt.*,ad.name as reviewer_name FROM major_transfer mt LEFT JOIN admin ad ON mt.reviewed_by=ad.admin_id WHERE mt.sno=%s ORDER BY mt.apply_date DESC",(sno,))
        abandons=q("SELECT sa.*,c.name as cn,ad.name as reviewer_name FROM score_abandon sa JOIN course c ON sa.cno=c.cno LEFT JOIN admin ad ON sa.reviewed_by=ad.admin_id WHERE sa.sno=%s ORDER BY sa.apply_date DESC",(sno,))
        return render_template('student_dashboard.html',stu=stu,scores=scores,rps=rps,transfers=transfers,abandons=abandons,sno=sno,gpa=gpa_val,avg_score=avg_score,avg_score_w=avg_score_w,total_credits=total_credits,college_stats=college_stats,major_stats=major_stats)

# =============================================================================
#                               管理员功能模块 (ADMIN)
# =============================================================================
# 学院管理：查看/新增/编辑/删除
@app.route('/admin/college')
@login_required
@role_required('admin')
def admin_college():
    colleges=q("SELECT * FROM college ORDER BY college_id")
    return render_template('admin_college.html',colleges=colleges)

@app.route('/admin/college/add',methods=['POST'])
@login_required
@role_required('admin')
def admin_college_add():
    n=request.form.get('name','').strip(); d=request.form.get('dean','')
    desc=request.form.get('description','')
    if n:
        try: e("INSERT INTO college(name,dean,description) VALUES(%s,%s,%s)",(n,d,desc)); flash('添加成功','success')
        except Exception as ex: flash(f'失败: {ex}','danger')
    return redirect(url_for('admin_college'))

@app.route('/admin/college/del/<int:cid>')
@login_required
@role_required('admin')
def admin_college_del(cid):
    try: e("DELETE FROM college WHERE college_id=%s",(cid,)); flash('已删除','info')
    except Exception as ex: flash(f'失败(可能有关联专业): {ex}','danger')
    return redirect(url_for('admin_college'))

@app.route('/admin/college/edit/<int:cid>',methods=['POST'])
@login_required
@role_required('admin')
def admin_college_edit(cid):
    n=request.form.get('name','').strip(); d=request.form.get('dean','')
    desc=request.form.get('description','')
    if n:
        try: e("UPDATE college SET name=%s,dean=%s,description=%s WHERE college_id=%s",(n,d,desc,cid)); flash('修改成功','success')
        except Exception as ex: flash(f'失败: {ex}','danger')
    return redirect(url_for('admin_college'))

@app.route('/admin/major')
@login_required
@role_required('admin')
def admin_major():
    majors=q("SELECT m.major_id,m.name,c.name as cn,m.college_id FROM major m JOIN college c ON m.college_id=c.college_id ORDER BY m.major_id")
    colleges=q("SELECT * FROM college ORDER BY college_id")
    return render_template('admin_major.html',majors=majors,colleges=colleges)

@app.route('/admin/major/add',methods=['POST'])
@login_required
@role_required('admin')
def admin_major_add():
    n=request.form.get('name','').strip(); cid=request.form.get('college_id')
    if n and cid:
        try: e("INSERT INTO major(name,college_id) VALUES(%s,%s)",(n,cid)); flash('添加成功','success')
        except Exception as ex: flash(f'失败: {ex}','danger')
    return redirect(url_for('admin_major'))

@app.route('/admin/major/del/<int:mid>')
@login_required
@role_required('admin')
def admin_major_del(mid):
    try: e("DELETE FROM major WHERE major_id=%s",(mid,)); flash('已删除','info')
    except Exception as ex: flash(f'删除失败: {ex}','danger')
    return redirect(url_for('admin_major'))

@app.route('/admin/major/edit/<int:mid>',methods=['POST'])
@login_required
@role_required('admin')
def admin_major_edit(mid):
    n=request.form.get('name','').strip(); cid=request.form.get('college_id')
    if n and cid:
        try: e("UPDATE major SET name=%s,college_id=%s WHERE major_id=%s",(n,cid,mid)); flash('修改成功','success')
        except Exception as ex: flash(f'失败: {ex}','danger')
    return redirect(url_for('admin_major'))

@app.route('/admin/class')
@login_required
@role_required('admin')
def admin_class():
    classes=q("SELECT c.class_id,c.name,m.name as mn,c.enroll_year,c.major_id FROM class c JOIN major m ON c.major_id=m.major_id ORDER BY c.enroll_year, c.name")
    majors=q("SELECT * FROM major ORDER BY major_id")
    return render_template('admin_class.html',classes=classes,majors=majors)

@app.route('/admin/class/add',methods=['POST'])
@login_required
@role_required('admin')
def admin_class_add():
    n=request.form.get('name','').strip(); mid=request.form.get('major_id')
    ey=request.form.get('enroll_year') or None
    if n and mid:
        try: e("INSERT INTO class(name,major_id,enroll_year) VALUES(%s,%s,%s)",(n,mid,ey)); flash('添加成功','success')
        except Exception as ex: flash(f'失败: {ex}','danger')
    return redirect(url_for('admin_class'))

@app.route('/admin/class/del/<int:cid>')
@login_required
@role_required('admin')
def admin_class_del(cid):
    try: e("DELETE FROM class WHERE class_id=%s",(cid,)); flash('已删除','info')
    except Exception as ex: flash(f'失败: {ex}','danger')
    return redirect(url_for('admin_class'))

@app.route('/admin/class/edit/<int:cid>',methods=['POST'])
@login_required
@role_required('admin')
def admin_class_edit(cid):
    n=request.form.get('name','').strip(); mid=request.form.get('major_id')
    ey=request.form.get('enroll_year') or None
    if n and mid:
        try: e("UPDATE class SET name=%s,major_id=%s,enroll_year=%s WHERE class_id=%s",(n,mid,ey,cid)); flash('修改成功','success')
        except Exception as ex: flash(f'失败: {ex}','danger')
    return redirect(url_for('admin_class'))

@app.route('/admin/student')
@login_required
@role_required('admin')
def admin_student():
    students=q("SELECT s.*,c.name as cn FROM student s LEFT JOIN class c ON s.class_id=c.class_id ORDER BY s.sno")
    classes=q("SELECT * FROM class ORDER BY class_id")
    return render_template('admin_student.html',students=students,classes=classes)

@app.route('/admin/student/add',methods=['POST'])
@login_required
@role_required('admin')
def admin_student_add():
    sno=request.form.get('sno','').strip(); name=request.form.get('name','').strip()
    g=request.form.get('gender','M'); b=request.form.get('birth','') or None
    st=request.form.get('status','active'); ey=request.form.get('enroll_year') or None
    ph=request.form.get('phone','') or None; em=request.form.get('email','') or None
    cid=request.form.get('class_id') or None
    if sno and name:
        try: e("INSERT INTO student(sno,name,gender,birth,status,enroll_year,phone,email,class_id) VALUES(%s,%s,%s,%s,%s,%s,%s,%s,%s)",(sno,name,g,b,st,ey,ph,em,cid)); flash('添加成功','success')
        except Exception as ex: flash(f'失败: {ex}','danger')
    return redirect(url_for('admin_student'))

@app.route('/admin/student/del/<sno>')
@login_required
@role_required('admin')
def admin_student_del(sno):
    try: e("DELETE FROM student WHERE sno=%s",(sno,)); flash('已删除','info')
    except Exception as ex: flash(f'失败: {ex}','danger')
    return redirect(url_for('admin_student'))

@app.route('/admin/student/edit/<sno>',methods=['POST'])
@login_required
@role_required('admin')
def admin_student_edit(sno):
    name=request.form.get('name','').strip(); g=request.form.get('gender','M')
    b=request.form.get('birth','') or None; st=request.form.get('status','active')
    ey=request.form.get('enroll_year') or None; ph=request.form.get('phone','') or None
    em=request.form.get('email','') or None; cid=request.form.get('class_id') or None
    if name:
        try: e("UPDATE student SET name=%s,gender=%s,birth=%s,status=%s,enroll_year=%s,phone=%s,email=%s,class_id=%s WHERE sno=%s",
               (name,g,b,st,ey,ph,em,cid,sno)); flash('修改成功','success')
        except Exception as ex: flash(f'失败: {ex}','danger')
    return redirect(url_for('admin_student'))

@app.route('/admin/api/student/<sno>/courses')
@login_required
@role_required('admin')
def admin_student_courses_api(sno):
    """返回学生的课程列表和可选课程"""
    courses=q("SELECT sc.cno,c.name,c.credit,c.semester,c.type,sc.score,sc.status FROM sc JOIN course c ON sc.cno=c.cno WHERE sc.sno=%s ORDER BY c.semester,c.name",(sno,))
    # All courses not taken by this student
    available=q("SELECT c.cno,c.name,c.credit,c.semester,c.type FROM course c WHERE c.cno NOT IN (SELECT cno FROM sc WHERE sno=%s) ORDER BY c.semester,c.name",(sno,))
    stu=q("SELECT sno,name FROM student WHERE sno=%s",(sno,),one=True)
    result = {
        'sno': sno, 'sname': stu[1] if stu else '',
        'courses': [{'cno':r[0],'name':r[1],'credit':r[2],'semester':r[3],'type':r[4],'score':r[5],'status':r[6]} for r in courses],
        'available': [{'cno':r[0],'name':r[1],'credit':r[2],'semester':r[3],'type':r[4]} for r in available]
    }
    return jsonify(result)

@app.route('/admin/student/<sno>/enroll',methods=['POST'])
@login_required
@role_required('admin')
def admin_student_enroll(sno):
    cno=request.form.get('cno')
    if sno and cno:
        try: e("INSERT INTO sc(sno,cno) VALUES(%s,%s)",(sno,cno)); flash('置课成功','success')
        except Exception as ex: flash(f'失败: {ex}','danger')
    return redirect(url_for('admin_student'))

@app.route('/admin/student/<sno>/drop/<int:cno>')
@login_required
@role_required('admin')
def admin_student_drop(sno,cno):
    e("DELETE FROM sc WHERE sno=%s AND cno=%s",(sno,cno)); flash('退课成功','info')
    return redirect(url_for('admin_student'))

@app.route('/admin/teacher')
@login_required
@role_required('admin')
def admin_teacher():
    teachers=q("SELECT t.*,c.name as cn FROM teacher t LEFT JOIN college c ON t.college_id=c.college_id ORDER BY t.tno")
    colleges=q("SELECT * FROM college ORDER BY college_id")
    return render_template('admin_teacher.html',teachers=teachers,colleges=colleges)

@app.route('/admin/teacher/add',methods=['POST'])
@login_required
@role_required('admin')
def admin_teacher_add():
    n=request.form.get('name','').strip(); g=request.form.get('gender','M'); t=request.form.get('title','')
    cid=request.form.get('college_id') or None
    if n:
        try: e("INSERT INTO teacher(name,gender,title,college_id) VALUES(%s,%s,%s,%s)",(n,g,t,cid)); flash('添加成功','success')
        except Exception as ex: flash(f'失败: {ex}','danger')
    return redirect(url_for('admin_teacher'))

@app.route('/admin/teacher/del/<int:tno>')
@login_required
@role_required('admin')
def admin_teacher_del(tno):
    try: e("DELETE FROM teacher WHERE tno=%s",(tno,)); flash('已删除','info')
    except Exception as ex: flash(f'失败: {ex}','danger')
    return redirect(url_for('admin_teacher'))

@app.route('/admin/teacher/edit/<int:tno>',methods=['POST'])
@login_required
@role_required('admin')
def admin_teacher_edit(tno):
    n=request.form.get('name','').strip(); g=request.form.get('gender','M')
    title=request.form.get('title','') or None; cid=request.form.get('college_id') or None
    if n:
        try: e("UPDATE teacher SET name=%s,gender=%s,title=%s,college_id=%s WHERE tno=%s",(n,g,title,cid,tno)); flash('修改成功','success')
        except Exception as ex: flash(f'失败: {ex}','danger')
    return redirect(url_for('admin_teacher'))

@app.route('/admin/admin_mgmt')
@login_required
@role_required('admin')
def admin_admin_mgmt():
    admins=q("SELECT * FROM admin ORDER BY admin_id")
    return render_template('admin_admin.html',admins=admins)

@app.route('/admin/admin_mgmt/add',methods=['POST'])
@login_required
@role_required('admin')
def admin_admin_add():
    jn=request.form.get('job_no','').strip(); n=request.form.get('name','').strip()
    g=request.form.get('gender','M'); ph=request.form.get('phone','') or None
    em=request.form.get('email','') or None; dp=request.form.get('department','') or None
    t=request.form.get('title','') or None
    if jn and n:
        try: e("INSERT INTO admin(job_no,name,gender,phone,email,department,title) VALUES(%s,%s,%s,%s,%s,%s,%s)",(jn,n,g,ph,em,dp,t)); flash('添加成功','success')
        except Exception as ex: flash(f'失败: {ex}','danger')
    return redirect(url_for('admin_admin_mgmt'))

@app.route('/admin/admin_mgmt/del/<int:aid>')
@login_required
@role_required('admin')
def admin_admin_del(aid):
    try: e("DELETE FROM admin WHERE admin_id=%s",(aid,)); flash('已删除','info')
    except Exception as ex: flash(f'删除失败(可能有关联账号): {ex}','danger')
    return redirect(url_for('admin_admin_mgmt'))

@app.route('/admin/admin_mgmt/edit/<int:aid>',methods=['POST'])
@login_required
@role_required('admin')
def admin_admin_edit(aid):
    jn=request.form.get('job_no','').strip(); n=request.form.get('name','').strip()
    g=request.form.get('gender','M'); ph=request.form.get('phone','') or None
    em=request.form.get('email','') or None; dp=request.form.get('department','') or None
    t=request.form.get('title','') or None
    if jn and n:
        try: e("UPDATE admin SET job_no=%s,name=%s,gender=%s,phone=%s,email=%s,department=%s,title=%s WHERE admin_id=%s",(jn,n,g,ph,em,dp,t,aid)); flash('修改成功','success')
        except Exception as ex: flash(f'失败: {ex}','danger')
    return redirect(url_for('admin_admin_mgmt'))

@app.route('/admin/course')
@login_required
@role_required('admin')
def admin_course():
    courses=q("SELECT c.*,t.name as tn FROM course c LEFT JOIN teacher t ON c.tno=t.tno ORDER BY c.semester, FIELD(c.type,'通修','专业必修','专业选修','自由选修'), c.name")
    teachers=q("SELECT * FROM teacher ORDER BY tno")
    return render_template('admin_course.html',courses=courses,teachers=teachers)

@app.route('/admin/course/add',methods=['POST'])
@login_required
@role_required('admin')
def admin_course_add():
    n=request.form.get('name','').strip(); cr=request.form.get('credit',0,type=int)
    h=request.form.get('hours',0,type=int) or None; tp=request.form.get('type','专业必修')
    tno=request.form.get('tno') or None; sem=request.form.get('semester','2025-2026春')
    if n and cr>0:
        try: e("INSERT INTO course(name,credit,hours,type,tno,semester) VALUES(%s,%s,%s,%s,%s,%s)",(n,cr,h,tp,tno,sem)); flash('添加成功','success')
        except Exception as ex: flash(f'失败: {ex}','danger')
    return redirect(url_for('admin_course'))

@app.route('/admin/course/del/<int:cno>')
@login_required
@role_required('admin')
def admin_course_del(cno):
    try: e("DELETE FROM course WHERE cno=%s",(cno,)); flash('已删除','info')
    except Exception as ex: flash(f'失败: {ex}','danger')
    return redirect(url_for('admin_course'))

@app.route('/admin/course/edit/<int:cno>',methods=['POST'])
@login_required
@role_required('admin')
def admin_course_edit(cno):
    n=request.form.get('name','').strip(); cr=request.form.get('credit',0,type=int)
    h=request.form.get('hours',0,type=int) or None; tp=request.form.get('type','专业必修')
    tno=request.form.get('tno') or None; sem=request.form.get('semester','2025-2026春')
    if n and cr>0:
        try: e("UPDATE course SET name=%s,credit=%s,hours=%s,type=%s,tno=%s,semester=%s WHERE cno=%s",(n,cr,h,tp,tno,sem,cno)); flash('修改成功','success')
        except Exception as ex: flash(f'失败: {ex}','danger')
    return redirect(url_for('admin_course'))

@app.route('/admin/score')
@login_required
@role_required('admin')
def admin_score():
    courses=q("SELECT c.cno,c.name,c.credit,c.semester,COUNT(sc.sno) as cnt,ROUND(AVG(sc.score),1) as avg FROM course c LEFT JOIN sc ON c.cno=sc.cno AND " \
    "sc.status='normal' GROUP BY c.cno,c.name,c.credit,c.semester ORDER BY c.semester DESC, c.cno")
    course_scores={}
    for c in courses:
        cno=c[0]
        ss=q("SELECT sc.sno,st.name,sc.score,sc.status FROM sc JOIN student st ON sc.sno=st.sno WHERE sc.cno=%s ORDER BY st.sno",(cno,))
        course_scores[cno]=ss
    students=q("SELECT sno,name FROM student ORDER BY sno")
    all_courses=q("SELECT cno,name FROM course ORDER BY cno")
    return render_template('admin_score.html',courses=courses,course_scores=course_scores,students=students,all_courses=all_courses)

@app.route('/admin/score/update/<sno>/<int:cno>',methods=['POST'])
@login_required
@role_required('admin')
def admin_score_update(sno,cno):
    sc=request.form.get('score','')
    try:
        if sc=='': e("UPDATE sc SET score=NULL WHERE sno=%s AND cno=%s",(sno,cno))
        else:
            v=float(sc)
            if v<0 or v>100: flash('成绩0-100','danger')
            else: e("UPDATE sc SET score=%s WHERE sno=%s AND cno=%s",(v,sno,cno)); flash('录入成功','success')
    except: flash('无效数字','danger')
    return redirect(url_for('admin_score'))

# ===== 事务1：选课（调用存储过程 sp_enroll_student，含 FOR UPDATE 行锁）=====
@app.route('/admin/score/enroll',methods=['POST'])
@login_required
@role_required('admin')
def admin_score_enroll():
    sno=request.form.get('sno',''); cno=request.form.get('cno')
    if sno and cno:
        c=db(); cur=c.cursor()                             # ← 事务开始
        try:
            cur.callproc('sp_enroll_student',(sno,cno))    # 存储过程内部验证+INSERT+COMMIT
            c.commit()                                      # ← 提交
            flash('选课成功','success')
        except Exception as ex: c.rollback(); flash(f'失败: {ex}','danger')
        finally: c.close()
    return redirect(url_for('admin_score'))

@app.route('/admin/score/del/<sno>/<int:cno>')
@login_required
@role_required('admin')
def admin_score_del(sno,cno):
    try: e("DELETE FROM sc WHERE sno=%s AND cno=%s",(sno,cno)); flash('退课成功','info')
    except Exception as ex: flash(f'失败: {ex}','danger')
    return redirect(url_for('admin_score'))

@app.route('/admin/approvals')
@login_required
@role_required('admin')
def admin_approvals():
    transfers=q("SELECT mt.*,s.name as sn,c1.name as fc,c2.name as tc,ad.name as reviewer_name FROM major_transfer mt JOIN student s ON mt.sno=s.sno JOIN class c1 ON mt.from_class_id=c1.class_id JOIN class c2 ON mt.to_class_id=c2.class_id LEFT JOIN admin ad ON mt.reviewed_by=ad.admin_id ORDER BY FIELD(mt.status,'pending','approved','rejected'), mt.apply_date DESC")
    abandons=q("SELECT sa.*,s.name as sn,c.name as cn,ad.name as reviewer_name FROM score_abandon sa JOIN student s ON sa.sno=s.sno JOIN course c ON sa.cno=c.cno LEFT JOIN admin ad ON sa.reviewed_by=ad.admin_id ORDER BY FIELD(sa.status,'pending','approved','rejected'), sa.apply_date DESC")
    return render_template('admin_approvals.html',transfers=transfers,abandons=abandons)

# ===== 事务2：审批转专业（调用存储过程 sp_approve_transfer，双表原子更新）=====
@app.route('/admin/transfer/approve/<int:tid>',methods=['POST'])
@login_required
@role_required('admin')
def admin_transfer_approve(tid):
    action=request.form.get('action'); comment=request.form.get('comment','')
    c=db(); cur=c.cursor()                             # ← 事务开始
    try:
        cur.callproc('sp_approve_transfer',(tid,action,comment,session['ref_id']))
        c.commit()                                       # ← 提交：两张表一起生效
        flash('审批完成','success')
    except Exception as ex: c.rollback(); flash(f'失败: {ex}','danger')
    finally: c.close()
    return redirect(url_for('admin_approvals'))

# ===== 事务3：审批放弃成绩（调用存储过程 sp_approve_abandon，双表原子更新）=====
@app.route('/admin/abandon/approve/<int:aid>',methods=['POST'])
@login_required
@role_required('admin')
def admin_abandon_approve(aid):
    action=request.form.get('action'); comment=request.form.get('comment','')
    c=db(); cur=c.cursor()                             # ← 事务开始
    try:
        cur.callproc('sp_approve_abandon',(aid,action,comment,session['ref_id']))
        c.commit()                                       # ← 提交
        flash('审批完成','success')
    except Exception as ex: c.rollback(); flash(f'失败: {ex}','danger')
    finally: c.close()
    return redirect(url_for('admin_approvals'))

@app.route('/admin/rp')
@login_required
@role_required('admin')
def admin_rp():
    rps=q("SELECT rp.*,s.name as sn,ad.name as creator_name FROM reward_punishment rp JOIN student s ON rp.sno=s.sno " \
    "LEFT JOIN admin ad ON rp.created_by=ad.admin_id ORDER BY rp.rp_date DESC")
    students=q("SELECT sno,name FROM student ORDER BY sno")
    return render_template('admin_rp.html',rps=rps,students=students)

@app.route('/admin/rp/add',methods=['POST'])
@login_required
@role_required('admin')
def admin_rp_add():
    sno=request.form.get('sno'); tp=request.form.get('type'); title=request.form.get('title','').strip()
    desc=request.form.get('description',''); dt=request.form.get('rp_date')
    if sno and tp and title and dt: e("INSERT INTO reward_punishment(sno,type,title,description,rp_date,created_by) " \
    "VALUES(%s,%s,%s,%s,%s,%s)",(sno,tp,title,desc,dt,session['ref_id'])); flash('添加成功','success')
    return redirect(url_for('admin_rp'))

@app.route('/admin/rp/del/<int:rid>')
@login_required
@role_required('admin')
def admin_rp_del(rid):
    e("UPDATE reward_punishment SET status='canceled' WHERE rp_id=%s",(rid,)); flash('已撤销','info')
    return redirect(url_for('admin_rp'))

@app.route('/admin/account')
@login_required
@role_required('admin')
def admin_account():
    accounts=q("""SELECT a.*,
        CASE WHEN a.role='admin' THEN ad.name
             WHEN a.role='teacher' THEN t.name
             WHEN a.role='student' THEN s.name
        END as ref_name
        FROM account a
        LEFT JOIN admin ad ON a.ref_id=ad.admin_id AND a.role='admin'
        LEFT JOIN teacher t ON a.ref_id=t.tno AND a.role='teacher'
        LEFT JOIN student s ON a.username=s.sno AND a.role='student'
        ORDER BY a.account_id""")
    admins=q("SELECT admin_id,job_no,name FROM admin ORDER BY admin_id")
    teachers=q("SELECT tno,name FROM teacher ORDER BY tno")
    students=q("SELECT sno,name FROM student ORDER BY sno")
    return render_template('admin_account.html',accounts=accounts,admins=admins,teachers=teachers,students=students)

@app.route('/admin/account/add',methods=['POST'])
@login_required
@role_required('admin')
def admin_account_add():
    u=request.form.get('username','').strip(); p=request.form.get('password','')
    r=request.form.get('role'); ref=request.form.get('ref_id') or None
    if u and p and r:
        pw=hashlib.sha256(p.encode()).hexdigest()
        try: e("INSERT INTO account(username,password_hash,role,ref_id) VALUES(%s,%s,%s,%s)",(u,pw,r,ref)); flash('创建成功','success')
        except Exception as ex: flash(f'失败: {ex}','danger')
    return redirect(url_for('admin_account'))

@app.route('/admin/account/reset/<int:aid>')
@login_required
@role_required('admin')
def admin_account_reset(aid):
    pw=hashlib.sha256('123456'.encode()).hexdigest()
    try: e("UPDATE account SET password_hash=%s WHERE account_id=%s",(pw,aid)); flash('密码已重置为 123456','success')
    except Exception as ex: flash(f'失败: {ex}','danger')
    return redirect(url_for('admin_account'))

# ---------- 文件导入（Excel → 数据库）----------
# 支持从 .xlsx 批量导入学生和教师，自动创建账号；依赖 openpyxl
import openpyxl
from io import BytesIO

@app.route('/admin/import')
@login_required
@role_required('admin')
def admin_import():
    return render_template('admin_import.html')

@app.route('/admin/import/students',methods=['POST'])
@login_required
@role_required('admin')
def admin_import_students():
    file = request.files.get('file')
    if not file or not file.filename:
        flash('请选择文件','danger'); return redirect(url_for('admin_import'))
    try:
        wb = openpyxl.load_workbook(BytesIO(file.read()))
        ws = wb.active
        rows = list(ws.iter_rows(min_row=2, values_only=True))  # skip header
        ok = 0; skip = 0; err = 0
        c=db(); cur=c.cursor()
        for row in rows:
            if not row[0] or not row[1]: skip += 1; continue
            sno = str(row[0]).strip(); name = str(row[1]).strip()
            gender = str(row[2]).strip().upper() if row[2] else 'M'
            if gender not in ('M','F'): gender = 'M'
            birth = str(row[3]).strip()[:10] if row[3] else None
            enroll_year = int(row[4]) if row[4] else None
            phone = str(row[5]).strip() if row[5] else None
            email = str(row[6]).strip() if row[6] else None
            class_name = str(row[7]).strip() if row[7] else None
            class_id = None
            if class_name:
                cls = q("SELECT class_id FROM class WHERE name=%s",(class_name,),one=True)
                if cls: class_id = cls[0]
            try:
                cur.execute("INSERT IGNORE INTO student(sno,name,gender,birth,status,enroll_year,phone,email,class_id) " \
                "VALUES(%s,%s,%s,%s,'active',%s,%s,%s,%s)",
                  (sno,name,gender,birth,enroll_year,phone,email,class_id))
                if cur.rowcount > 0:
                    pw = hashlib.sha256(sno.encode()).hexdigest()
                    cur.execute("INSERT IGNORE INTO account(username,password_hash,role) " \
                    "VALUES(%s,%s,'student')",(sno,pw))
                    ok += 1
                else:
                    skip += 1
            except Exception: err += 1
        c.commit(); c.close()
        flash(f'导入完成：成功 {ok} 人，跳过 {skip} 行，失败 {err} 行','success' if ok>0 else 'warning')
    except Exception as ex:
        flash(f'文件解析失败: {ex}','danger')
    return redirect(url_for('admin_import'))

@app.route('/admin/import/teachers',methods=['POST'])
@login_required
@role_required('admin')
def admin_import_teachers():
    file = request.files.get('file')
    if not file or not file.filename:
        flash('请选择文件','danger'); return redirect(url_for('admin_import'))
    try:
        wb = openpyxl.load_workbook(BytesIO(file.read()))
        ws = wb.active
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        ok = 0; skip = 0; err = 0
        c=db(); cur=c.cursor()
        for row in rows:
            if not row[0] or not row[1]: skip += 1; continue
            tno_str = str(row[0]).strip(); name = str(row[1]).strip()
            gender = str(row[2]).strip().upper() if row[2] else 'M'
            if gender not in ('M','F'): gender = 'M'
            title = str(row[3]).strip() if row[3] else None
            college_name = str(row[4]).strip() if row[4] else None
            college_id = None
            if college_name:
                col = q("SELECT college_id FROM college WHERE name=%s",(college_name,),one=True)
                if col: college_id = col[0]
            try:
                cur.execute("INSERT IGNORE INTO teacher(tno,name,gender,title,college_id) VALUES(%s,%s,%s,%s,%s)",
                  (tno_str,name,gender,title,college_id))
                if cur.rowcount > 0:
                    pw = hashlib.sha256('123456'.encode()).hexdigest()
                    cur.execute("INSERT IGNORE INTO account(username,password_hash,role) VALUES(%s,%s,'teacher')",(f't_{tno_str}',pw))
                    ok += 1
                else:
                    skip += 1
            except Exception: err += 1
        c.commit(); c.close()
        flash(f'导入完成：成功 {ok} 人，跳过 {skip} 行，失败 {err} 行','success' if ok>0 else 'warning')
    except Exception as ex:
        flash(f'文件解析失败: {ex}','danger')
    return redirect(url_for('admin_import'))

@app.route('/admin/import/template/<tp>')
@login_required
@role_required('admin')
def admin_import_template(tp):
    wb = openpyxl.Workbook()
    ws = wb.active
    if tp == 'students':
        ws.append(['学号','姓名','性别(M/F)','出生日期(YYYY-MM-DD)','入学年份','电话','邮箱','班级名称'])
        ws.append(['PB25110001','张小明','M','2007-09-01',2025,'13800001111','zhangxm@mail.ustc.edu.cn','计科2501'])
        filename = '学生导入模板.xlsx'
    else:
        ws.append(['工号','姓名','性别(M/F)','职称','学院名称'])
        ws.append(['23','张教授','M','教授','计算机科学与技术学院'])
        filename = '教师导入模板.xlsx'
    # Set column widths
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 18
    output = BytesIO()
    wb.save(output); output.seek(0)
    from flask import send_file
    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name=filename)

# ---------- 管理员视频管理（查看/删除全部课程视频）----------
@app.route('/admin/videos')
@login_required
@role_required('admin')
def admin_videos():
    videos = q("""
        SELECT v.*, c.name as cn, t.name as tn FROM course_video v 
        JOIN course c ON v.course_id=c.cno 
        LEFT JOIN teacher t ON v.uploaded_by=t.tno
        ORDER BY v.upload_time DESC
    """)
    return render_template('admin_videos.html', videos=videos)

@app.route('/admin/videos/del/<int:vid>')
@login_required
@role_required('admin')
def admin_video_del(vid):
    v = q("SELECT file_path FROM course_video WHERE video_id=%s",(vid,),one=True)
    if v:
        fp = os.path.join(app.root_path,'uploads','videos',v[0])
        if os.path.exists(fp): os.remove(fp)
        e("DELETE FROM course_video WHERE video_id=%s",(vid,))
        flash('视频已删除','info')
    return redirect(url_for('admin_videos'))

# =============================================================================
#                               教师功能模块 (TEACHER)
# =============================================================================
# 查看授课课程学生名单、录入/修改成绩、成绩分布统计
@app.route('/teacher/score/<int:cno>')
@login_required
@role_required('teacher')
def teacher_score(cno):
    tno=session['ref_id']
    ck=q("SELECT * FROM course WHERE cno=%s AND tno=%s",(cno,tno),one=True)
    if not ck: flash('无权访问此课程','danger'); return redirect(url_for('index'))
    scores=q("SELECT sc.sno,st.name,st.class_id,c2.name as cn,sc.score,sc.status,st.phone,st.email " \
    "FROM sc JOIN student st ON sc.sno=st.sno LEFT JOIN class c2 ON st.class_id=c2.class_id WHERE sc.cno=%s ORDER BY sc.score IS NULL DESC, st.sno",(cno,))
    # 调用存储过程 sp_teacher_course_stats 获取统计
    c2=db(); cur2=c2.cursor()
    cur2.callproc('sp_teacher_course_stats',(cno,tno,0,0,0,0,0,0,0))
    cur2.execute("SELECT @_sp_teacher_course_stats_2, @_sp_teacher_course_stats_3, @_sp_teacher_course_stats_4, "
    "@_sp_teacher_course_stats_5, @_sp_teacher_course_stats_6, @_sp_teacher_course_stats_7, @_sp_teacher_course_stats_8")
    stats_r=cur2.fetchone(); c2.close()
    stats=(stats_r[0],stats_r[1],stats_r[2],stats_r[3],stats_r[4],stats_r[5],stats_r[6])
    dist=q("SELECT CASE WHEN score>=90 THEN '90-100' WHEN score>=80 THEN '80-89' WHEN score>=70 THEN '70-79' " \
    "WHEN score>=60 THEN '60-69' ELSE '<60' END as rng, COUNT(*) as cnt FROM sc WHERE cno=%s AND score IS NOT NULL AND status='normal' GROUP BY rng ORDER BY MIN(score) DESC",(cno,))
    return render_template('teacher_score.html',scores=scores,course=ck,stats=stats,dist=dist)

@app.route('/teacher/score/update/<int:cno>/<sno>',methods=['POST'])
@login_required
@role_required('teacher')
def teacher_score_update(cno,sno):
    tno=session['ref_id']
    ck=q("SELECT * FROM course WHERE cno=%s AND tno=%s",(cno,tno),one=True)
    if not ck: flash('无权','danger'); return redirect(url_for('index'))
    sc=request.form.get('score','')
    try:
        if sc=='': e("UPDATE sc SET score=NULL WHERE sno=%s AND cno=%s",(sno,cno))
        else:
            v=float(sc)
            if v<0 or v>100: flash('0-100','danger')
            else: e("UPDATE sc SET score=%s WHERE sno=%s AND cno=%s",(v,sno,cno)); flash('录入成功','success')
    except: flash('无效','danger')
    return redirect(url_for('teacher_score',cno=cno))

# ---------- 教师视频管理（为所授课程上传/删除视频）----------
@app.route('/teacher/videos')
@login_required
@role_required('teacher')
def teacher_videos():
    tno = session['ref_id']
    courses = q("SELECT c.cno,c.name,c.semester FROM course c WHERE c.tno=%s ORDER BY c.semester DESC",(tno,))
    videos = q("""
        SELECT v.*, c.name as cn FROM course_video v 
        JOIN course c ON v.course_id=c.cno 
        WHERE v.uploaded_by=%s ORDER BY v.upload_time DESC
    """,(tno,))
    return render_template('teacher_videos.html', courses=courses, videos=videos)

@app.route('/teacher/videos/upload', methods=['POST'])
@login_required
@role_required('teacher')
def teacher_video_upload():
    tno = session['ref_id']
    cno = request.form.get('course_id')
    title = request.form.get('title','').strip()
    desc = request.form.get('description','')
    file = request.files.get('video_file')
    if not cno or not title or not file or not file.filename:
        flash('请填写标题并选择视频文件','danger'); return redirect(url_for('teacher_videos'))
    # Verify course belongs to this teacher
    ck = q("SELECT * FROM course WHERE cno=%s AND tno=%s",(cno,tno),one=True)
    if not ck: flash('无权为此课程上传','danger'); return redirect(url_for('teacher_videos'))
    import uuid
    ext = file.filename.rsplit('.',1)[-1] if '.' in file.filename else 'mp4'
    fname = f"video_{uuid.uuid4().hex}.{ext}"
    filepath = os.path.join(app.root_path,'uploads','videos',fname)
    file.save(filepath)
    fsize = os.path.getsize(filepath)
    e("INSERT INTO course_video(course_id,title,description,file_path,file_size,uploaded_by) VALUES(%s,%s,%s,%s,%s,%s)",
      (cno,title,desc,fname,fsize,tno))
    flash('视频上传成功','success')
    return redirect(url_for('teacher_videos'))

@app.route('/teacher/videos/del/<int:vid>')
@login_required
@role_required('teacher')
def teacher_video_del(vid):
    tno = session['ref_id']
    v = q("SELECT file_path FROM course_video WHERE video_id=%s AND uploaded_by=%s",(vid,tno),one=True)
    if v:
        fp = os.path.join(app.root_path,'uploads','videos',v[0])
        if os.path.exists(fp): os.remove(fp)
        e("DELETE FROM course_video WHERE video_id=%s",(vid,))
        flash('视频已删除','info')
    return redirect(url_for('teacher_videos'))

@app.route('/student/videos')
@login_required
@role_required('student')
def student_videos():
    sno = session['username']
    videos = q("""
        SELECT v.*, c.name as cn, c.semester FROM course_video v 
        JOIN course c ON v.course_id=c.cno 
        JOIN sc ON sc.cno=c.cno AND sc.sno=%s
        WHERE v.course_id=sc.cno
        ORDER BY c.semester DESC, v.upload_time DESC
    """,(sno,))
    return render_template('student_videos.html', videos=videos)

# =============================================================================
#                               学生功能模块 (STUDENT)
# =============================================================================
# 转专业申请、放弃成绩申请、选课/退课、奖项申请（含文件上传）、个人信息编辑
@app.route('/student/transfer',methods=['POST'])
@login_required
@role_required('student')
def student_transfer():
    sno=session['username']; fcid=request.form.get('from_class_id'); tcid=request.form.get('to_class_id')
    reason=request.form.get('reason','')
    p=q("SELECT * FROM major_transfer WHERE sno=%s AND status='pending'",(sno,),one=True)
    if p: flash('已有待审批的转专业申请','warning'); return redirect(url_for('index'))
    if fcid and tcid:
        try: e("INSERT INTO major_transfer(sno,from_class_id,to_class_id,reason) " \
        "VALUES(%s,%s,%s,%s)",(sno,fcid,tcid,reason)); flash('申请已提交','success')
        except Exception as ex: flash(f'失败: {ex}','danger')
    return redirect(url_for('index'))

@app.route('/student/abandon',methods=['POST'])
@login_required
@role_required('student')
def student_abandon():
    sno=session['username']; cno=request.form.get('cno'); reason=request.form.get('reason','')
    cnt=q("SELECT COUNT(*) FROM score_abandon WHERE sno=%s AND status='approved'",(sno,),one=True)
    if cnt and cnt[0]>=2: flash('已达最大放弃次数(2次)','warning'); return redirect(url_for('index'))
    if cno:
        try: e("INSERT INTO score_abandon(sno,cno,reason) VALUES(%s,%s,%s)",(sno,cno,reason)); flash('申请已提交','success')
        except Exception as ex: flash(f'失败: {ex}','danger')
    return redirect(url_for('index'))

@app.route('/student/course/enroll',methods=['POST'])
@login_required
@role_required('student')
def student_enroll():
    sno=session['username']; cno=request.form.get('cno')
    if cno:
        # Check semester is open for enrollment
        sem=q("SELECT semester FROM course WHERE cno=%s",(cno,),one=True)
        if sem and sem[0] not in open_enroll_semesters():
            flash('该学期课程当前不可选课','danger'); return redirect(url_for('student_enroll_page'))
        try: e("INSERT INTO sc(sno,cno) VALUES(%s,%s)",(sno,cno)); flash('选课成功','success')
        except Exception as ex: flash(f'失败: {ex}','danger')
    return redirect(url_for('student_enroll_page'))

@app.route('/student/course/drop/<int:cno>')
@login_required
@role_required('student')
def student_drop(cno):
    sno=session['username']
    sem=q("SELECT semester FROM course WHERE cno=%s",(cno,),one=True)
    if not sem:
        flash('课程不存在','danger')
    elif sem[0] not in open_enroll_semesters():
        flash('该学期课程当前不可退课','danger')
    else:
        try: e("DELETE FROM sc WHERE sno=%s AND cno=%s",(sno,cno)); flash('退课成功','info')
        except Exception as ex: flash(f'失败: {ex}','danger')
    return redirect(url_for('student_enroll_page'))

# Student: Award application with file upload
@app.route('/student/award/apply',methods=['POST'])
@login_required
@role_required('student')
def student_award_apply():
    sno=session['username']
    title=request.form.get('title','').strip()
    desc=request.form.get('description','')
    file=request.files.get('cert_file')
    file_path=None
    if file and file.filename:
        import uuid
        ext=file.filename.rsplit('.',1)[-1] if '.' in file.filename else 'pdf'
        fname=f"{uuid.uuid4().hex}.{ext}"
        file.save(os.path.join(app.root_path,'uploads','awards',fname))
        file_path=fname
    if title:
        try: e("INSERT INTO award_application(sno,title,description,file_path) VALUES(%s,%s,%s,%s)",(sno,title,desc,file_path)); flash('奖项申请已提交','success')
        except Exception as ex: flash(f'失败: {ex}','danger')
    else: flash('请填写奖项名称','warning')
    return redirect(url_for('student_apply'))

# Student: Profile (view + edit phone/email/photo)
@app.route('/student/profile',methods=['GET','POST'])
@login_required
@role_required('student')
def student_profile():
    sno=session['username']
    if request.method=='POST':
        phone=request.form.get('phone','').strip() or None
        email=request.form.get('email','').strip() or None
        photo=request.files.get('photo')
        if photo and photo.filename:
            # 删除旧照片
            old=q("SELECT photo_path FROM student WHERE sno=%s",(sno,),one=True)
            if old and old[0]:
                old_fp=os.path.join(app.root_path,'uploads','photos',old[0])
                if os.path.exists(old_fp): os.remove(old_fp)
            import uuid
            ext=photo.filename.rsplit('.',1)[-1] if '.' in photo.filename else 'jpg'
            fname=f"photo_{uuid.uuid4().hex}.{ext}"
            photo.save(os.path.join(app.root_path,'uploads','photos',fname))
            e("UPDATE student SET phone=%s,email=%s,photo_path=%s WHERE sno=%s",(phone,email,fname,sno))
        else:
            e("UPDATE student SET phone=%s,email=%s WHERE sno=%s",(phone,email,sno))
        flash('个人信息已更新','success')
        return redirect(url_for('student_profile'))
    stu=q("SELECT s.*,c.name as cn FROM student s LEFT JOIN class c ON s.class_id=c.class_id WHERE s.sno=%s",(sno,),one=True)
    return render_template('student_profile.html',stu=stu)

# Admin: Award management
@app.route('/admin/awards')
@login_required
@role_required('admin')
def admin_awards():
    awards=q("SELECT a.*,s.name as sn FROM award_application a JOIN student s ON a.sno=s.sno ORDER BY a.apply_date DESC")
    return render_template('admin_awards.html',awards=awards)

# 事务4：审批奖项（触发器 trg_award_to_reward 自动同步奖惩）=====
@app.route('/admin/award/approve/<int:aid>',methods=['POST'])
@login_required
@role_required('admin')
def admin_award_approve(aid):
    action=request.form.get('action'); comment=request.form.get('comment','')
    if action=='approve':
        e("UPDATE award_application SET status='approved',review_date=NOW(),review_comment=%s,reviewed_by=%s WHERE app_id=%s",(comment,session['ref_id'],aid))
        flash('已批准','success')
    else:
        e("UPDATE award_application SET status='rejected',review_date=NOW(),review_comment=%s,reviewed_by=%s WHERE app_id=%s",(comment,session['ref_id'],aid))
        flash('已拒绝','info')
    return redirect(url_for('admin_awards'))

# ---------- 上传文件访问（图片/视频/文档等静态资源）----------
@app.route('/uploads/<path:fname>')
@login_required
def uploaded_file(fname):
    """提供 uploads 目录下文件的访问，需登录"""
    return send_from_directory(os.path.join(app.root_path,'uploads'),fname)

# ---------- 学生成绩页（按学期分组，含GPA/排名/学分卡片）----------
@app.route('/student/scores')
@login_required
@role_required('student')
def student_scores():
    sno=session['username']
    scores=q("SELECT sc.cno,c.name,c.credit,sc.score,sc.status,fn_score_to_gp_43(sc.score) as gp,c.semester,c.type " \
    "FROM sc JOIN course c ON sc.cno=c.cno WHERE sc.sno=%s ORDER BY c.semester, c.name",(sno,))
    # 调用存储过程 sp_student_gpa_rank（GPA + 均分 + 学分 + 学院总人数 + 排名）
    c=db(); cur=c.cursor()
    cur.callproc('sp_student_gpa_rank',(sno,0,0,0,0,0,0))
    cur.execute("SELECT @_sp_student_gpa_rank_1, @_sp_student_gpa_rank_2, @_sp_student_gpa_rank_3, "
    "@_sp_student_gpa_rank_4, @_sp_student_gpa_rank_5, @_sp_student_gpa_rank_6")
    gpa_r=cur.fetchone()
    college_total = gpa_r[4] or 0; college_rank = gpa_r[5] or '-'
    # 专业排名（同专业同年级）
    cur.callproc('sp_student_major_rank',(sno,0,0))
    cur.execute("SELECT @_sp_student_major_rank_1, @_sp_student_major_rank_2")
    mr=cur.fetchone(); c.close()
    major_total = mr[0] or 0; major_rank = mr[1] or '-'
    return render_template('student_scores.html',scores=scores,sno=sno,gpa=gpa_r[0],avg_score=gpa_r[1],avg_score_w=gpa_r[2],
                           total_credits=gpa_r[3],college_total=college_total,college_rank=college_rank,major_total=major_total,major_rank=major_rank)

# ---------- 学生奖惩记录页 ----------
@app.route('/student/rp')
@login_required
@role_required('student')
def student_rp():
    sno=session['username']
    rps=q("SELECT rp.*,ad.name as creator_name FROM reward_punishment rp LEFT JOIN admin ad ON rp.created_by=ad.admin_id WHERE rp.sno=%s ORDER BY rp.rp_date DESC",(sno,))
    return render_template('student_rp.html',rps=rps,sno=sno)

# ---------- 学生申请中心（转专业 + 放弃成绩 + 奖项申请）----------
@app.route('/student/apply')
@login_required
@role_required('student')
def student_apply():
    sno=session['username']
    stu=q("SELECT s.*,c.name as cn FROM student s LEFT JOIN class c ON s.class_id=c.class_id WHERE s.sno=%s",(sno,),one=True)
    scores=q("SELECT sc.cno,c.name,c.credit,sc.score,sc.status FROM sc JOIN course c ON sc.cno=c.cno WHERE sc.sno=%s ORDER BY c.name",(sno,))
    transfers=q("SELECT mt.*,ad.name as reviewer_name FROM major_transfer mt LEFT JOIN admin ad ON mt.reviewed_by=ad.admin_id WHERE mt.sno=%s ORDER BY mt.apply_date DESC",(sno,))
    abandons=q("SELECT sa.*,c.name as cn,ad.name as reviewer_name FROM score_abandon sa JOIN course c ON sa.cno=c.cno LEFT JOIN admin ad ON sa.reviewed_by=ad.admin_id WHERE sa.sno=%s ORDER BY sa.apply_date DESC",(sno,))
    awards=q("SELECT * FROM award_application WHERE sno=%s ORDER BY apply_date DESC",(sno,))
    # 同级班级（与当前学生同入学年份的其他班级）
    same_year_classes=q("""SELECT c.class_id, c.name, m.name as mn FROM class c JOIN major m ON c.major_id=m.major_id 
        WHERE c.enroll_year=(SELECT c2.enroll_year FROM student s2 JOIN class c2 ON s2.class_id=c2.class_id WHERE s2.sno=%s)
        ORDER BY c.name""",(sno,))
    return render_template('student_apply.html',stu=stu,scores=scores,transfers=transfers,abandons=abandons,awards=awards,classes=same_year_classes,sno=sno)

# ---------- 学生选课中心（按学期显示可选课程）----------
@app.route('/student/enroll')
@login_required
@role_required('student')
def student_enroll_page():
    sno=session['username']
    open_sems = open_enroll_semesters()
    scores=q("SELECT cno FROM sc WHERE sno=%s",(sno,))
    enrolled = {s[0] for s in scores}
    # Only show courses from open semesters
    placeholders = ','.join(['%s']*len(open_sems))
    all_courses=q(f"SELECT c.*, t.name as tn FROM course c LEFT JOIN teacher t ON c.tno=t.tno WHERE c.semester IN ({placeholders}) ORDER BY c.semester, c.cno", open_sems)
    available = [c for c in all_courses if c[0] not in enrolled]
    wk, phase = week_of_semester()
    return render_template('student_enroll.html',available=available,sno=sno,phase=phase,week=wk,open_sems=open_sems)

# ---------- 教师个人信息页 ----------
@app.route('/teacher/profile')
@login_required
@role_required('teacher')
def teacher_profile():
    tno=session['ref_id']
    teacher=q("SELECT * FROM teacher WHERE tno=%s",(tno,),one=True)
    course_count=q("SELECT COUNT(*) FROM course WHERE tno=%s",(tno,),one=True)
    student_count=q("SELECT COUNT(DISTINCT sc.sno) FROM sc JOIN course c ON sc.cno=c.cno WHERE c.tno=%s",(tno,),one=True)
    return render_template('teacher_profile.html',teacher=teacher,course_count=course_count[0],student_count=student_count[0])

# ---------- 管理员个人信息页 ----------
@app.route('/admin/profile')
@login_required
@role_required('admin')
def admin_profile():
    aid=session['ref_id']
    admin=q("SELECT * FROM admin WHERE admin_id=%s",(aid,),one=True) if aid else None
    return render_template('admin_profile.html',admin=admin)

# ============ 应用启动入口 ============
if __name__=='__main__':
    print('='*50)
    print('  学生管理系统')
    print('  http://127.0.0.1:5000')
    print('  admin/123456 | t_1/123456 | PB23111650/123456')
    print('='*50)
    app.run(debug=True,host='127.0.0.1',port=5000)
